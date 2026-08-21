"""Turn an uploaded CSV / JSON / XLSX export into headers plus string-valued rows.

Everything downstream (mapping, conversion) works on `dict[str, str]`, so the quirks of each
format - BOMs, sniffed delimiters, nested JSON, Excel cell types - are dealt with once, here.
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from datetime import date, datetime, time
from typing import Any

MAX_FILE_BYTES = 2 * 1024 * 1024
MAX_ROWS = 2_000
CSV_SUFFIXES = (".csv", ".tsv")
JSON_SUFFIXES = (".json",)
XLSX_SUFFIXES = (".xlsx", ".xlsm")
SUPPORTED_SUFFIXES = CSV_SUFFIXES + JSON_SUFFIXES + XLSX_SUFFIXES


class FileParseError(RuntimeError):
    """The upload could not be turned into a table; the caller answers 4xx."""


class UnsupportedFile(FileParseError):
    """Not a CSV, JSON or XLSX file -> 415."""


class EmptyFile(FileParseError):
    """Parsed fine but contains no data rows -> 422."""


class FileTooLarge(FileParseError):
    """Over the byte or row cap -> 413."""


@dataclass(slots=True)
class ParsedTable:
    """Column names in file order plus one dict per data row, values as trimmed strings."""

    headers: list[str]
    rows: list[dict[str, str]]
    filename: str = ""
    # Nested JSON keys that were flattened to dotted names, kept for the preview notes.
    flattened: list[str] = field(default_factory=list)


def _stringify(value: Any) -> str:
    """One cell -> text, keeping dates ISO and integral floats free of a trailing `.0`."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == time.min else value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else repr(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value).strip()


def _decode(data: bytes, filename: str) -> str:
    """UTF-8 with the BOM stripped; a decode failure is reported, never guessed around."""
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise FileParseError(f"{filename} is not valid UTF-8 text: {exc}") from exc


def _clean_headers(raw: list[Any]) -> list[str]:
    """Trim, name blanks, and suffix duplicates so no column is silently lost."""
    headers: list[str] = []
    for index, name in enumerate(raw, start=1):
        clean = (str(name) if name is not None else "").strip() or f"column_{index}"
        if clean in headers:
            clean = f"{clean} ({headers.count(clean) + 1})"
        headers.append(clean)
    return headers


def _is_blank(row: dict[str, str]) -> bool:
    return not any(value.strip() for value in row.values())


def _parse_csv(text: str, filename: str) -> ParsedTable:
    """Sniff the delimiter (falling back to a comma) and read every non-empty row."""
    sample = text[:8192]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = "\t" if filename.lower().endswith(".tsv") else ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        first = next(reader)
    except StopIteration as exc:
        raise EmptyFile(f"{filename} contains no rows.") from exc

    headers = _clean_headers(list(first))
    rows: list[dict[str, str]] = []
    for values in reader:
        row = {
            header: _stringify(values[i]) if i < len(values) else ""
            for i, header in enumerate(headers)
        }
        if not _is_blank(row):
            rows.append(row)
        if len(rows) > MAX_ROWS:
            raise FileTooLarge(f"{filename} has more than {MAX_ROWS} rows.")
    return ParsedTable(headers=headers, rows=rows, filename=filename)


def _flatten(obj: dict[str, Any], prefix: str = "") -> dict[str, Any]:
    """Nested objects become dotted keys (`vendor.name`); lists stay whole, later JSON-encoded."""
    flat: dict[str, Any] = {}
    for key, value in obj.items():
        path = f"{prefix}{key}"
        if isinstance(value, dict):
            flat.update(_flatten(value, f"{path}."))
        else:
            flat[path] = value
    return flat


def _json_records(payload: Any, filename: str) -> list[dict[str, Any]]:
    """An array of objects, an object with exactly one array value, or a single object = one row."""
    if isinstance(payload, list):
        records = payload
    elif isinstance(payload, dict):
        arrays = [value for value in payload.values() if isinstance(value, list)]
        records = arrays[0] if len(arrays) == 1 else [payload]
    else:
        raise FileParseError(
            f"{filename} must contain a JSON object or array, not {type(payload).__name__}."
        )

    if not all(isinstance(record, dict) for record in records):
        raise FileParseError(f"{filename} must contain JSON objects; found a non-object entry.")
    return records


def _parse_json(text: str, filename: str) -> ParsedTable:
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise FileParseError(f"{filename} is not valid JSON: {exc}") from exc

    records = _json_records(payload, filename)
    if len(records) > MAX_ROWS:
        raise FileTooLarge(f"{filename} has more than {MAX_ROWS} rows.")

    headers: list[str] = []
    flattened: list[str] = []
    rows: list[dict[str, str]] = []
    for record in records:
        flat = _flatten(record)
        for key in flat:
            if key not in headers:
                headers.append(key)
            if "." in key and key not in flattened:
                flattened.append(key)
        row = {key: _stringify(value) for key, value in flat.items()}
        if not _is_blank(row):
            rows.append(row)

    # Give every row every column so the mapper always sees a rectangular table.
    rows = [{header: row.get(header, "") for header in headers} for row in rows]
    return ParsedTable(headers=headers, rows=rows, filename=filename, flattened=flattened)


def _parse_xlsx(data: bytes, filename: str) -> ParsedTable:
    """First sheet, first non-empty row as headers, cached values only (`data_only`)."""
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except FileTooLarge:
        raise
    except Exception as exc:
        raise FileParseError(f"Could not read {filename} as a workbook: {exc}") from exc

    headers: list[str] = []
    rows: list[dict[str, str]] = []
    try:
        sheet = workbook.worksheets[0]
        for values in sheet.iter_rows(values_only=True):
            cells = [_stringify(value) for value in values]
            if not any(cell.strip() for cell in cells):
                continue
            if not headers:
                headers = _clean_headers(list(cells))
                continue
            rows.append(
                {header: cells[i] if i < len(cells) else "" for i, header in enumerate(headers)}
            )
            if len(rows) > MAX_ROWS:
                raise FileTooLarge(f"{filename} has more than {MAX_ROWS} rows.")
    finally:
        workbook.close()

    if not headers:
        raise EmptyFile(f"{filename} has no header row.")
    return ParsedTable(headers=headers, rows=rows, filename=filename)


def parse_tabular(filename: str, data: bytes) -> ParsedTable:
    """Dispatch on the file extension. Raises a `FileParseError` subclass; never returns empty."""
    if len(data) > MAX_FILE_BYTES:
        raise FileTooLarge(f"{filename} is {len(data)} bytes; the limit is {MAX_FILE_BYTES} bytes.")

    name = (filename or "").lower()
    if name.endswith(CSV_SUFFIXES):
        table = _parse_csv(_decode(data, filename), filename)
    elif name.endswith(JSON_SUFFIXES):
        table = _parse_json(_decode(data, filename), filename)
    elif name.endswith(XLSX_SUFFIXES):
        table = _parse_xlsx(data, filename)
    else:
        raise UnsupportedFile(
            f"Unsupported file type {filename!r}. Upload one of: {', '.join(SUPPORTED_SUFFIXES)}."
        )

    if not table.rows:
        raise EmptyFile(f"{filename} contains no data rows.")
    return table
